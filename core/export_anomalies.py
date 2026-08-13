# -*- coding: utf-8 -*-
"""
Export Excel consolidant TOUTES les anomalies (OT + Avis) detectees sur le
perimetre filtre courant (periode / poste / atelier / division), avec pour
chaque ligne : le(s) KPI en anomalie, le(s) responsable(s) et l'(les)
action(s) recommandee(s).

Deux feuilles :
 - "Anomalies OT" : UNE ligne par OT (pas par KPI). Si un OT est en
   anomalie sur plusieurs KPI, toutes les anomalies sont listees dans la
   colonne "Anomalie" (separees par " | "), pas de duplication de l'OT.
 - "Anomalies Avis" : idem, une ligne par Avis.

CORRIGE : la version precedente faisait un simple pd.concat() des
DataFrames d'anomalies par KPI, donc un OT en anomalie sur 3 KPI
apparaissait 3 fois (une ligne par KPI). Desormais, on consolide par
cle (Ordre pour les OT, Avis pour les avis) : une seule ligne par
element, avec le nombre d'anomalies et le detail dans une colonne.
"""
import io
import pandas as pd

OT_COLS = [
    "Avis", "Ordre", "Désignation", "Numéro d'objet", "Poste technique",
    "Désignation du poste technique", "Équipement",
    "Description de l'objet technique", "Zone de tri", "Poste travail princ.",
    "Groupe gestionnaires", "Divis. planification", "Statut système",
    "Statut utilisateur", "Saisi par", "Date de début planifiée", "Créé le",
]

AVIS_COLS = [
    "Avis", "Ordre", "Description", "Poste technique",
    "Désignation du poste technique", "Poste travail princ.",
    "Divis. planification", "Groupe gestionnaires", "Statut système",
    "Statut utilisateur", "Type de travail", "Type d'avis", "Priorité",
    "Texte de la priorité", "Type de priorité", "Auteur de l'avis",
    "Créé par", "Créé le",
]

AVIS_ONLY_KPI = "Taux d'approbation des Avis"

def _extract_cols(df: pd.DataFrame, cols: list) -> pd.DataFrame:
    """Garde uniquement les colonnes demandees, dans l'ordre, en ajoutant
    des colonnes vides si absentes de df (extraction source variable)."""
    out = pd.DataFrame(index=df.index)
    for c in cols:
        out[c] = df[c] if c in df.columns else ""
    return out

def _consolidate_by_key(df_long: pd.DataFrame, key_col: str, static_cols: list) -> pd.DataFrame:
    """Regroupe les lignes en double (même OT/Avis apparu pour plusieurs
    KPI en anomalie) en UNE SEULE ligne par clé (Ordre ou Avis).
    - Les colonnes "statiques" (identité de l'OT/Avis) gardent leur
      première valeur rencontrée (identiques d'une ligne à l'autre pour
      une même clé).
    - "Anomalie" / "Responsable" / "Action recommandée" sont fusionnées :
      toutes les valeurs uniques, séparées par " | ".
    - Ajoute une colonne "Nb anomalies" = nombre de KPI en anomalie pour
      cet OT/Avis.
    """
    if df_long.empty or key_col not in df_long.columns:
        return df_long

    # Exclure les lignes sans clé exploitable (Ordre/Avis vide) du
    # regroupement — elles restent telles quelles, chacune sa ligne.
    has_key = df_long[key_col].notna() & (df_long[key_col].astype(str).str.strip() != "")
    df_keyed = df_long[has_key].copy()
    df_nokey = df_long[~has_key].copy()

    if df_keyed.empty:
        return df_long

    nb_anom = df_keyed.groupby(key_col)["Anomalie"].transform("count")
    df_keyed["Nb anomalies"] = nb_anom

    agg_dict = {c: "first" for c in static_cols if c in df_keyed.columns and c != key_col}
    agg_dict["Nb anomalies"] = "first"
    agg_dict["Anomalie"] = lambda s: " | ".join(dict.fromkeys(s.astype(str)))  # valeurs uniques, ordre conservé
    agg_dict["Responsable"] = lambda s: " | ".join(dict.fromkeys(s.astype(str)))
    agg_dict["Action recommandée"] = lambda s: " | ".join(dict.fromkeys(s.astype(str)))

    grouped = df_keyed.groupby(key_col, as_index=False, sort=False).agg(agg_dict)

    if not df_nokey.empty:
        df_nokey["Nb anomalies"] = 1
        grouped = pd.concat([grouped, df_nokey], ignore_index=True)

    return grouped

def build_anomalies_workbook(anomaly_dfs: dict, kpi_resp_map: dict,
                              act_map: dict) -> bytes:
    """
    Construit un classeur Excel (2 feuilles) a partir de anomaly_dfs
    (deja filtre sur periode/poste/atelier/division cote appelant).
    Retourne les bytes du fichier .xlsx.
    """
    ot_rows = []
    avis_rows = []

    for kpi_name, df_anom in anomaly_dfs.items():
        if df_anom is None or df_anom.empty:
            continue

        responsable = kpi_resp_map.get(kpi_name, "Non assigné")
        action = act_map.get(kpi_name, "")

        if kpi_name == AVIS_ONLY_KPI:
            sub = _extract_cols(df_anom, AVIS_COLS)
            sub["Anomalie"] = kpi_name
            sub["Responsable"] = responsable
            sub["Action recommandée"] = action
            avis_rows.append(sub)
        else:
            sub = _extract_cols(df_anom, OT_COLS)
            sub["Anomalie"] = kpi_name
            sub["Responsable"] = responsable
            sub["Action recommandée"] = action
            ot_rows.append(sub)

    df_ot_long = (
        pd.concat(ot_rows, ignore_index=True) if ot_rows
        else pd.DataFrame(columns=OT_COLS + ["Anomalie", "Responsable", "Action recommandée"])
    )
    df_avis_long = (
        pd.concat(avis_rows, ignore_index=True) if avis_rows
        else pd.DataFrame(columns=AVIS_COLS + ["Anomalie", "Responsable", "Action recommandée"])
    )

    # CORRIGÉ : consolidation par clé -> une seule ligne par OT / par Avis,
    # avec toutes les anomalies listées dans une seule colonne.
    df_ot_final = _consolidate_by_key(df_ot_long, "Ordre", OT_COLS)
    df_avis_final = _consolidate_by_key(df_avis_long, "Avis", AVIS_COLS)

    # Réordonner les colonnes : identité OT/Avis, puis Nb anomalies,
    # Anomalie, Responsable, Action recommandée.
    def _reorder(df_final: pd.DataFrame, base_cols: list) -> pd.DataFrame:
        if df_final.empty:
            return df_final
        tail = ["Nb anomalies", "Anomalie", "Responsable", "Action recommandée"]
        ordered = [c for c in base_cols if c in df_final.columns] + [c for c in tail if c in df_final.columns]
        remaining = [c for c in df_final.columns if c not in ordered]
        return df_final[ordered + remaining]

    df_ot_final = _reorder(df_ot_final, OT_COLS)
    df_avis_final = _reorder(df_avis_final, AVIS_COLS)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_ot_final.to_excel(writer, sheet_name="Anomalies OT", index=False)
        df_avis_final.to_excel(writer, sheet_name="Anomalies Avis", index=False)

        # Mise en forme simple : entetes en gras + largeur auto approx.
        from openpyxl.styles import Font, PatternFill, Alignment
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        for sheet_name, df_sheet in [("Anomalies OT", df_ot_final), ("Anomalies Avis", df_avis_final)]:
            ws = writer.sheets[sheet_name]
            for col_idx, col_name in enumerate(df_sheet.columns, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center")
                width = min(45, max(12, len(str(col_name)) + 4))
                ws.column_dimensions[cell.column_letter].width = width

    buf.seek(0)
    return buf.getvalue()
