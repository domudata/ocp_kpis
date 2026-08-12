# -*- coding: utf-8 -*-
import os
import pandas as pd
from openpyxl import load_workbook

from core.constants import LOWER_BETTER

def load_historical_kpis(filepath: str) -> pd.DataFrame:
    """Charge TOUT l'historique enregistré (une feuille par date dans le
    classeur). Aucune limite de nombre de dates ici — si l'app n'affiche
    que 2 dates, c'est que le fichier kpis/indicateurs_kpis.xlsx local n'a
    que 2 feuilles (probablement parce que le disque de l'app est éphémère
    et que seules les versions committées sur GitHub survivent aux
    redémarrages — voir l'onglet "Suivi & Evolution" > Historique).

    CORRIGÉ : parse maintenant AUSSI les sections ANOMALIES PERFORMANCE /
    ANOMALIES QUALITE (avant, elles étaient lues puis explicitement
    ignorées : `section = None`). Elles sont taguées _section =
    "ano_perf" / "ano_qual", en parallèle de "perf" / "qual" pour les
    valeurs de KPI.
    """
    if not os.path.exists(filepath):
        return pd.DataFrame()
    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception:
        return pd.DataFrame()

    records = []
    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            rows_data = list(ws.iter_rows(values_only=True))
            section = None
            headers = None
            for row in rows_data:
                cell0 = str(row[0]).strip() if row[0] else ""
                up = cell0.upper()
                if "ANOMALIES PERFORMANCE" in up:
                    section = "ano_perf"; headers = None; continue
                elif "ANOMALIES QUALITE" in up:
                    section = "ano_qual"; headers = None; continue
                elif "INDICATEURS DE PERFORMANCE" in up:
                    section = "perf"; headers = None; continue
                elif "INDICATEURS DE QUALITE" in up:
                    section = "qual"; headers = None; continue
                if section and headers is None and cell0:
                    headers = [str(c).strip() if c else "" for c in row]; continue
                if section and headers and cell0 and cell0 not in ("Cible", "Total general", "Total", ""):
                    entry = {"Date": sheet_name}
                    for j, h in enumerate(headers):
                        if j < len(row):
                            entry[h] = row[j]
                    entry["_section"] = section
                    records.append(entry)
        except Exception:
            continue

    wb.close()
    if not records:
        return pd.DataFrame()

    df = pd.DataFrame(records)
    df["Date_parsed"] = pd.to_datetime(
        df["Date"].str.replace("-", "/"), format="%d/%m/%Y", errors="coerce"
    )
    return df.sort_values("Date_parsed").reset_index(drop=True)

def calculate_variations(hist_df: pd.DataFrame) -> pd.DataFrame:
    if hist_df.empty or "Date" not in hist_df.columns:
        return pd.DataFrame()

    dates = sorted(hist_df["Date"].unique())
    if len(dates) < 2:
        return pd.DataFrame()

    perf_df = hist_df[hist_df["_section"] == "perf"].copy()
    qual_df = hist_df[hist_df["_section"] == "qual"].copy()
    variations = []

    from core.constants import QK, PK

    for i in range(1, len(dates)):
        prev_date, curr_date = dates[i - 1], dates[i]

        prev_perf = perf_df[perf_df["Date"] == prev_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        curr_perf = perf_df[perf_df["Date"] == curr_date].set_index("Poste de travail") if "Poste de travail" in perf_df.columns else pd.DataFrame()
        prev_qual = qual_df[qual_df["Date"] == prev_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()
        curr_qual = qual_df[qual_df["Date"] == curr_date].set_index("Poste de travail") if "Poste de travail" in qual_df.columns else pd.DataFrame()

        for sec_name, prev_d, curr_d, kpi_list in [
            ("Performance", prev_perf, curr_perf, QK + ["Score Performance"]),
            ("Qualite", prev_qual, curr_qual, PK + ["Score Qualite"]),
        ]:
            for poste in set(prev_d.index) & set(curr_d.index):
                for kpi in kpi_list:
                    if kpi not in prev_d.columns or kpi not in curr_d.columns:
                        continue
                    try:
                        pv = float(prev_d.loc[poste, kpi])
                    except Exception:
                        continue
                    try:
                        cv = float(curr_d.loc[poste, kpi])
                    except Exception:
                        continue

                    diff = cv - pv
                    pct = (diff / pv * 100) if pv != 0 else (100 if cv != 0 else 0)

                    if abs(diff) <= 0.5:
                        trend = "stabilite"
                    elif diff > 0.5:
                        trend = "hausse"
                    else:
                        trend = "baisse"

                    if trend == "stabilite":
                        sens = "Stable"
                    elif (trend == "hausse" and kpi not in LOWER_BETTER) or \
                         (trend == "baisse" and kpi in LOWER_BETTER):
                        sens = "Amelioration"
                    else:
                        sens = "Degradation"

                    variations.append({
                        "Date precedente": prev_date, "Date actuelle": curr_date,
                        "Poste": poste, "Type": sec_name, "KPI": kpi,
                        "Valeur precedente": round(pv, 2), "Valeur actuelle": round(cv, 2),
                        "Ecart": round(diff, 2), "Ecart %": round(pct, 2),
                        "Tendance": trend, "Sens": sens,
                    })

    return pd.DataFrame(variations)

def generate_journal(var_df: pd.DataFrame) -> pd.DataFrame:
    if var_df.empty:
        return pd.DataFrame()
    j = var_df.copy()
    j["Significatif"] = j["Ecart %"].abs() >= 5
    j = j[j["Significatif"]].copy()
    return j.sort_values(["Date actuelle", "Ecart %"], ascending=[True, False])

def calculate_rankings(var_df: pd.DataFrame):
    if var_df.empty:
        return pd.DataFrame(), pd.DataFrame()
    scores = {}
    for poste in var_df["Poste"].unique():
        pv = var_df[var_df["Poste"] == poste].copy()
        scores[poste] = sum(
            (-r["Ecart %"] if r["KPI"] in LOWER_BETTER else r["Ecart %"])
            for _, r in pv.iterrows()
        )
    ranked = sorted(scores.items(), key=lambda x: x[1], reverse=True)
    return (
        pd.DataFrame(ranked[:5], columns=["Poste", "Score variation"]),
        pd.DataFrame(ranked[-5:][::-1], columns=["Poste", "Score variation"]),
    )

# ──────────────────────────────────────────────
# NOUVEAU : évaluation sur les N dernières valeurs enregistrées
# (pas seulement la comparaison entre les 2 dernières périodes)
# ──────────────────────────────────────────────

def get_recent_window(hist_df: pd.DataFrame, section: str, kpi: str, n: int = 5) -> pd.DataFrame:
    """Pivot Poste x Date restreint aux n dernières dates enregistrées,
    pour un KPI et une section donnés (section : "perf"/"qual"/"ano_perf"/"ano_qual")."""
    if hist_df.empty or kpi not in hist_df.columns:
        return pd.DataFrame()
    sub = hist_df[hist_df["_section"] == section].copy()
    if sub.empty or "Poste de travail" not in sub.columns:
        return pd.DataFrame()
    sub["Date_str"] = sub["Date_parsed"].dt.strftime("%d/%m/%Y")
    pv = sub.pivot_table(index="Poste de travail", columns="Date_str", values=kpi, aggfunc="first")
    ordered_cols = sorted(pv.columns, key=lambda c: pd.to_datetime(c, format="%d/%m/%Y", errors="coerce"))
    pv = pv[ordered_cols]
    return pv.iloc[:, -n:] if n else pv

def evaluate_trend_last_n(hist_df: pd.DataFrame, section: str, kpi_list: list, n: int = 5) -> pd.DataFrame:
    """Pour chaque poste et chaque KPI (ou KPI d'anomalies), évalue la
    tendance sur les n DERNIÈRES valeurs enregistrées — pas seulement les
    2 dernières comme calculate_variations(). Compare la première valeur
    de la fenêtre à la dernière.
    section : "perf"/"qual" pour les KPI, "ano_perf"/"ano_qual" pour les
    anomalies (dans ce cas, toujours "plus bas = mieux", cible implicite 0).
    """
    is_anomalie = section in ("ano_perf", "ano_qual")
    rows = []
    for kpi in kpi_list:
        pv = get_recent_window(hist_df, section, kpi, n)
        if pv.empty:
            continue
        for poste in pv.index:
            vals = pv.loc[poste].dropna().tolist()
            if len(vals) < 2:
                continue
            first, last = vals[0], vals[-1]
            diff = last - first
            pct = (diff / first * 100) if first else (100 if last else 0)
            lower = True if is_anomalie else (kpi in LOWER_BETTER)
            if abs(diff) < (0.5 if is_anomalie else 0) and abs(pct) < 2:
                tendance = "Stable"
            elif (diff > 0 and not lower) or (diff < 0 and lower):
                tendance = "Amélioration"
            elif diff == 0:
                tendance = "Stable"
            else:
                tendance = "Dégradation"
            rows.append({
                "Poste": poste, "KPI": kpi, "Nb valeurs": len(vals),
                "Valeurs": vals, "Première": round(first, 1), "Dernière": round(last, 1),
                "Écart": round(diff, 1), "Écart %": round(pct, 1), "Tendance": tendance,
            })
    return pd.DataFrame(rows)
