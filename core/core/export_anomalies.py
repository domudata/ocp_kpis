# -*- coding: utf-8 -*-
"""
Export Excel consolidant TOUTES les anomalies (OT + Avis) detectees sur le
perimetre filtre courant (periode / poste / atelier / division), avec pour
chaque ligne : le KPI en anomalie, le responsable et l'action recommandee.

Deux feuilles :
  - "Anomalies OT"   : une ligne par (OT, KPI en anomalie)
  - "Anomalies Avis" : une ligne par (Avis, KPI en anomalie)
"""
import io
import pandas as pd

OT_COLS = [
    "Avis", "Ordre", "Désignation", "Numéro d'objet", "Poste technique",
    "Désignation du poste technique", "Équipement",
    "Description de l'objet technique", "Zone de tri", "Poste travail princ.",
    "Groupe gestionnaires", "Divis. planification", "Statut système",
    "Statut utilisateur", "Saisi par", "Date de début planifiée", "Créé le",
]

AVIS_COLS = [
    "Avis", "Ordre", "Description", "Poste technique",
    "Désignation du poste technique", "Poste travail princ.",
    "Divis. planification", "Groupe gestionnaires", "Statut système",
    "Statut utilisateur", "Type de travail", "Type d'avis", "Priorité",
    "Texte de la priorité", "Type de priorité", "Auteur de l'avis",
    "Créé par", "Créé le",
]

AVIS_ONLY_KPI = "Taux d'approbation des Avis"


def _extract_cols(df: pd.DataFrame, cols: list) -> pd.DataFrame:
    """Garde uniquement les colonnes demandees, dans l'ordre, en ajoutant
    des colonnes vides si absentes de df (extraction source variable)."""
    out = pd.DataFrame(index=df.index)
    for c in cols:
        out[c] = df[c] if c in df.columns else ""
    return out


def build_anomalies_workbook(anomaly_dfs: dict, kpi_resp_map: dict,
                              act_map: dict) -> bytes:
    """
    Construit un classeur Excel (2 feuilles) a partir de anomaly_dfs
    (deja filtre sur periode/poste/atelier/division cote appelant).
    Retourne les bytes du fichier .xlsx.
    """
    ot_rows = []
    avis_rows = []

    for kpi_name, df_anom in anomaly_dfs.items():
        if df_anom is None or df_anom.empty:
            continue

        responsable = kpi_resp_map.get(kpi_name, "Non assigné")
        action = act_map.get(kpi_name, "")

        if kpi_name == AVIS_ONLY_KPI:
            sub = _extract_cols(df_anom, AVIS_COLS)
            sub["Anomalie"] = kpi_name
            sub["Responsable"] = responsable
            sub["Action recommandée"] = action
            avis_rows.append(sub)
        else:
            sub = _extract_cols(df_anom, OT_COLS)
            sub["Anomalie"] = kpi_name
            sub["Responsable"] = responsable
            sub["Action recommandée"] = action
            ot_rows.append(sub)

    df_ot_final = (
        pd.concat(ot_rows, ignore_index=True) if ot_rows
        else pd.DataFrame(columns=OT_COLS + ["Anomalie", "Responsable", "Action recommandée"])
    )
    df_avis_final = (
        pd.concat(avis_rows, ignore_index=True) if avis_rows
        else pd.DataFrame(columns=AVIS_COLS + ["Anomalie", "Responsable", "Action recommandée"])
    )

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_ot_final.to_excel(writer, sheet_name="Anomalies OT", index=False)
        df_avis_final.to_excel(writer, sheet_name="Anomalies Avis", index=False)

        # Mise en forme simple : entetes en gras + largeur auto approx.
        from openpyxl.styles import Font, PatternFill, Alignment
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        for sheet_name, df_sheet in [("Anomalies OT", df_ot_final), ("Anomalies Avis", df_avis_final)]:
            ws = writer.sheets[sheet_name]
            for col_idx, col_name in enumerate(df_sheet.columns, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center")
                width = min(45, max(12, len(str(col_name)) + 4))
                ws.column_dimensions[cell.column_letter].width = width

    buf.seek(0)
    return buf.getvalue()
