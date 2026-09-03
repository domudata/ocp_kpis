# -*- coding: utf-8 -*-
import io
import os

import pandas as pd
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def save_kpis_to_excel(prows, pcols, qrows, qcols,
                        ano_p_r, ano_p_c, ano_q_r, ano_q_c,
                        sheet_name: str) -> None:
    kpis_dir = "kpis"

    hf = Font(bold=True, color="FFFFFF", size=10)
    hfl = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    tf = Font(bold=True, size=12, color="1E3A5F")
    tb = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # CORRIGÉ : chaque étape à risque (création dossier, écriture disque)
    # est maintenant tracée avec un message explicite dans la sidebar en
    # cas d'échec, au lieu d'un `except: pass` qui avalait silencieusement
    # toute erreur (permissions, disque en lecture seule sur Streamlit
    # Cloud, etc.) — l'app continuait à tourner normalement sans que la
    # date ne soit réellement sauvegardée, sans aucun signal visible.
    try:
        os.makedirs(kpis_dir, exist_ok=True)
    except Exception as e:
        st.sidebar.error(f"❌ Impossible de créer le dossier '{kpis_dir}/' : {e}")
        return

    filepath = os.path.join(kpis_dir, "indicateurs_kpis.xlsx")

    # ── Synchronisation GitHub-first (avec diagnostic détaillé) ──────────
    # Avant de charger le fichier local, on tente de récupérer la
    # DERNIÈRE version publiée sur GitHub et on l'utilise comme point de
    # départ si elle est plus complète que ce qu'il y a en local.
    _diag = []  # trace complète, affichée en fin de fonction pour diagnostic
    try:
        from core.github_publish import download_file as _gh_download, is_configured as _gh_is_configured
        if _gh_is_configured():
            _diag.append("GitHub configuré : oui")
            _remote_bytes, _dl_err = _gh_download("kpis/indicateurs_kpis.xlsx")
            if _remote_bytes:
                _remote_wb_check = load_workbook(io.BytesIO(_remote_bytes), read_only=True)
                _diag.append(f"Distant GitHub : {len(_remote_wb_check.sheetnames)} feuille(s) — {_remote_wb_check.sheetnames}")
                _remote_wb_check.close()
                _use_remote = True
                if os.path.exists(filepath):
                    try:
                        _local_wb = load_workbook(filepath, read_only=True)
                        _local_sheets = _local_wb.sheetnames
                        _diag.append(f"Local avant sync : {len(_local_sheets)} feuille(s) — {_local_sheets}")
                        _remote_wb = load_workbook(io.BytesIO(_remote_bytes), read_only=True)
                        _use_remote = len(_remote_wb.sheetnames) >= len(_local_wb.sheetnames)
                        _local_wb.close()
                        _remote_wb.close()
                    except Exception as _e2:
                        _diag.append(f"Erreur lecture local pour comparaison : {_e2}")
                        _use_remote = True
                else:
                    _diag.append("Local avant sync : fichier absent")
                _diag.append(f"Décision : {'utiliser le distant' if _use_remote else 'garder le local'}")
                if _use_remote:
                    with open(filepath, "wb") as _f:
                        _f.write(_remote_bytes)
            elif _dl_err:
                _diag.append(f"Téléchargement distant : échec — {_dl_err}")
            else:
                _diag.append("Téléchargement distant : fichier inexistant sur GitHub (normal si 1ère fois)")
        else:
            _diag.append("GitHub configuré : NON (secrets absents) — historique local uniquement, non persistant entre redémarrages")
    except Exception as _sync_e:
        _diag.append(f"Synchronisation GitHub : exception — {_sync_e}")

    sn = (
        str(sheet_name)
        .replace("/", "-").replace("\\", "-").replace("*", "")
        .replace("?", "").replace("[", "").replace("]", "")[:31]
    )
    _diag.append(f"Nom de feuille pour cette date : '{sn}'")

    try:
        wb = load_workbook(filepath)
        _diag.append(f"Fichier local chargé : {len(wb.sheetnames)} feuille(s) — {wb.sheetnames}")
    except FileNotFoundError:
        wb = Workbook()
        _diag.append("Fichier local absent → nouveau classeur vide créé")
    except Exception as e:
        st.sidebar.error(f"❌ Impossible d'ouvrir '{filepath}' (fichier corrompu ?) : {e}")
        return

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    if sn in wb.sheetnames:
        del wb[sn]
        _diag.append(f"Feuille '{sn}' existait déjà → supprimée avant recréation (mise à jour de la même date)")

    ws = wb.create_sheet(sn)

    def ws_sec(title, cols, rows, sr):
        ws.cell(row=sr, column=1, value=title).font = tf
        sr += 1
        for j, c in enumerate(cols, 1):
            cl = ws.cell(row=sr, column=j, value=c)
            cl.font = hf
            cl.fill = hfl
            cl.alignment = Alignment(horizontal='center')
            cl.border = tb
        sr += 1
        for r in rows:
            for j, c in enumerate(cols, 1):
                cl = ws.cell(row=sr, column=j, value=r.get(c, ""))
                cl.border = tb
                cl.alignment = Alignment(horizontal='center')
            sr += 1
        return sr + 1

    rn = 1
    rn = ws_sec("INDICATEURS DE PERFORMANCE", pcols, prows, rn)
    if ano_p_c and ano_p_r:
        rn = ws_sec("ANOMALIES PERFORMANCE", ano_p_c, ano_p_r, rn)
    rn = ws_sec("INDICATEURS DE QUALITE", qcols, qrows, rn)
    if ano_q_c and ano_q_r:
        rn = ws_sec("ANOMALIES QUALITE", ano_q_c, ano_q_r, rn)

    _diag.append(f"Après ajout de '{sn}' : {len(wb.sheetnames)} feuille(s) — {wb.sheetnames}")

    try:
        wb.save(filepath)
    except Exception as e:
        st.sidebar.error(
            f"❌ Échec de la sauvegarde de l'historique ('{filepath}') : {e}\n\n"
            f"La date '{sn}' n'a PAS été enregistrée. Téléchargez quand même "
            f"le fichier via le bouton de secours ci-dessous si besoin."
        )
        try:
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            st.sidebar.download_button(
                "⬇️ Télécharger quand même (sauvegarde disque échouée)",
                data=buf, file_name="indicateurs_kpis.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"fallback_dl_{sn}",
            )
        except Exception:
            pass
        return

    st.sidebar.success(f"✅ Historique mis à jour : date '{sn}' enregistrée dans {filepath} ({len(wb.sheetnames)} date(s) au total)")
    with st.sidebar.expander("🔍 Diagnostic détaillé historique", expanded=False):
        for line in _diag:
            st.caption(line)

def export_btn(df: pd.DataFrame, filename: str) -> None:
    buf = io.BytesIO()
    df.to_excel(buf, index=False, engine='openpyxl')
    buf.seek(0)
    st.download_button(
        "📥 Exporter Excel", data=buf,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
